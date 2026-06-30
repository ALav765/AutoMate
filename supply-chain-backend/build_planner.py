import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

import math
import os

# ─────────────────────────────────────────────
# FILE PATHS (overridable via env vars set by runner.py)
# ─────────────────────────────────────────────
FORECAST_FILE = os.environ.get("FORECAST_FILE", "/mnt/project/TEMPLATE_Forecast_SOH_1.xlsx")
BOM_FILE      = os.environ.get("BOM_FILE",      "/mnt/project/TEMPLATE_BOM_1.xlsx")
COMP_SOH_FILE = os.environ.get("COMP_SOH_FILE", "/mnt/project/TEMPLATE_BOM_Component_SOH_2.xlsx")
PRICES_FILE   = os.environ.get("PRICES_FILE",   "/mnt/project/TEMPLATE_BOM_Prices.xlsx")
VENDOR_FILE   = os.environ.get("VENDOR_FILE",   "/mnt/project/TEMPLATE_Vendor_Master_1.xlsx")
OUTPUT_FILE   = os.environ.get("OUTPUT_FILE",   "/mnt/user-data/outputs/Integrated_Supply_Plan.xlsx")

# ─────────────────────────────────────────────
# COLOR PALETTE
# ─────────────────────────────────────────────
C_NAVY      = "1B3A6B"
C_BLUE      = "2563EB"
C_LTBLUE    = "DBEAFE"
C_TEAL      = "0F766E"
C_LTGREEN   = "D1FAE5"
C_AMBER     = "F59E0B"
C_LTAMBER   = "FEF3C7"
C_RED       = "DC2626"
C_LTRED     = "FEE2E2"
C_WHITE     = "FFFFFF"
C_GRAY1     = "F1F5F9"
C_GRAY2     = "E2E8F0"
C_GRAY3     = "94A3B8"
C_BLACK     = "0F172A"
C_PURPLE    = "7C3AED"
C_LTPURPLE  = "EDE9FE"

MONTHS = ["Jun-25","Jul-25","Aug-25","Sep-25"]
MONTH_DAYS = {"Jun-25":30,"Jul-25":31,"Aug-25":31,"Sep-25":30}

def fill(hex_color, fill_type="solid"):
    return PatternFill(fill_type=fill_type, fgColor=hex_color)

def font(bold=False, color=C_BLACK, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides="all"):
    s = Side(style="thin", color=C_GRAY2)
    m = Side(style="medium", color=C_GRAY3)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    if sides == "bottom_medium":
        return Border(bottom=m)
    if sides == "top_medium":
        return Border(top=m)
    return Border()

def num_fmt(ws, cell, fmt="#,##0"):
    ws[cell].number_format = fmt

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def apply_header(ws, row, cols_data, bg=C_NAVY, fg=C_WHITE, sz=10, bold=True):
    for col_idx, val in enumerate(cols_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = fill(bg)
        cell.font = font(bold=bold, color=fg, size=sz)
        cell.alignment = align(h="center", wrap=True)
        cell.border = thin_border()

def apply_subheader(ws, row, cols_data, bg=C_BLUE, fg=C_WHITE, sz=9):
    for col_idx, val in enumerate(cols_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = fill(bg)
        cell.font = font(bold=True, color=fg, size=sz)
        cell.alignment = align(h="center", wrap=True)
        cell.border = thin_border()

def write_cell(ws, row, col, value, bg=None, bold=False, color=C_BLACK,
               num_format=None, h="right", sz=9, italic=False, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = fill(bg)
    cell.font = font(bold=bold, color=color, size=sz)
    cell.alignment = align(h=h, wrap=wrap)
    cell.border = thin_border()
    if num_format:
        cell.number_format = num_format
    return cell

# ─────────────────────────────────────────────
# LOAD DATA
# ─────────────────────────────────────────────
print("Loading data...")

df_fc = pd.read_excel(FORECAST_FILE, header=None, skiprows=4)
df_fc.columns = ['Category','_b','SAP_Code','Product_Name','SOH','Jun','Jul','Aug','Sep']
df_fc = df_fc[df_fc['SAP_Code'].apply(
    lambda x: str(x).strip().replace('.','').isdigit() if pd.notna(x) else False)].copy()
df_fc['SAP_Code'] = df_fc['SAP_Code'].apply(lambda x: str(int(float(x))))
for col in ['SOH','Jun','Jul','Aug','Sep']:
    df_fc[col] = pd.to_numeric(df_fc[col], errors='coerce').fillna(0)
df_fc = df_fc.reset_index(drop=True)

# BOM
df_bom_raw = pd.read_excel(BOM_FILE, header=None, skiprows=2)
df_bom_raw.columns = ['FG_SAP','FG_Name','Comp_Code','Comp_Cat','Lot_Size',
                       'Comp_Desc','_g','Qty_per_FG','Lead_Time','_j','Source','Supplier_Code']
df_bom = df_bom_raw[
    df_bom_raw['FG_SAP'].apply(lambda x: str(x).strip().replace('.','').isdigit() if pd.notna(x) else False) &
    df_bom_raw['Source'].notna() &
    (df_bom_raw['Source'].astype(str).str.strip().isin(['Domestic','Import']))
].copy()
df_bom['FG_SAP'] = df_bom['FG_SAP'].apply(lambda x: str(int(float(x))))
df_bom['Comp_Code'] = df_bom['Comp_Code'].astype(str).str.strip()
df_bom['Qty_per_FG'] = pd.to_numeric(df_bom['Qty_per_FG'], errors='coerce').fillna(0)
df_bom['Lot_Size'] = pd.to_numeric(df_bom['Lot_Size'], errors='coerce').fillna(1)
df_bom['Lead_Time'] = pd.to_numeric(df_bom['Lead_Time'], errors='coerce').fillna(30)
# De-duplicate: first row per FG+Component
df_bom = df_bom.drop_duplicates(subset=['FG_SAP','Comp_Code'], keep='first')

# Component SOH
df_csoh_raw = pd.read_excel(COMP_SOH_FILE, header=None, skiprows=2)
df_csoh_raw.columns = ['Comp_Code','Comp_Desc','Category','SOH','UoM','Target_Days']
df_csoh = df_csoh_raw[df_csoh_raw['Comp_Code'].apply(
    lambda x: str(x).strip().replace('.','').isdigit() if pd.notna(x) else False)].copy()
df_csoh['Comp_Code'] = df_csoh['Comp_Code'].apply(lambda x: str(int(float(x))))
df_csoh['SOH'] = pd.to_numeric(df_csoh['SOH'], errors='coerce').fillna(0)
df_csoh['Target_Days'] = pd.to_numeric(df_csoh['Target_Days'], errors='coerce').fillna(0)
df_csoh = df_csoh.drop_duplicates(subset=['Comp_Code'], keep='first')

# Prices
df_prices_raw = pd.read_excel(PRICES_FILE, header=None, skiprows=2)
df_prices_raw.columns = ['Comp_Code','Comp_Desc','Category','Unit_Price','Currency','ABC_Class','Supplier_Code']
df_prices = df_prices_raw[df_prices_raw['Comp_Code'].apply(
    lambda x: str(x).strip().replace('.','').isdigit() if pd.notna(x) else False)].copy()
df_prices['Comp_Code'] = df_prices['Comp_Code'].apply(lambda x: str(int(float(x))))
df_prices['Unit_Price'] = pd.to_numeric(df_prices['Unit_Price'], errors='coerce').fillna(0)
df_prices = df_prices.drop_duplicates(subset=['Comp_Code'], keep='first')

# Vendor
df_vendor = pd.read_excel(VENDOR_FILE, header=None, skiprows=2, sheet_name='Vendor_Master')
df_vendor.columns = ['Supplier Code','Supplier Name','Contact Person','Email','Country','Dom_Import']
df_vendor['Supplier Code'] = df_vendor['Supplier Code'].astype(str).str.strip()

print(f"  Forecast SKUs: {len(df_fc)}")
print(f"  BOM rows: {len(df_bom)}")
print(f"  Components in SOH: {len(df_csoh)}")
print(f"  Components with prices: {len(df_prices)}")

# ─────────────────────────────────────────────
# FG INVENTORY TARGETS (days) by Category
# Based on typical FMCG practice - since no separate sheet provided
# ─────────────────────────────────────────────
FG_INV_TARGET = {
    'HIT Aerosol': 15, 'HIT Non Stop': 15, 'HIT Mat': 10, 'HIT Liquid & pump': 15,
    'HIT Magic Paper': 10, 'HIT Expert Piramida': 15, 'HIT Anti Roach': 15,
    'HIT Expert Electric Racket': 20, 'Baby Toiletries': 15, 'Baby Wipes': 15,
    'Stella Aerosol': 15, 'Matic Refill': 10, 'Matic Device': 20, 'Pocket': 10,
    'Solid': 10, 'Spray': 15, 'Car Vent': 20, 'Home Wardrobe': 15, 'Elec.Liq.': 15,
    'Car Hanging': 20, 'Naturalist': 20, 'Biosol': 10, 'Cap Gajah Lem Lalat': 15,
    'Cap Gajah Lem Perangkap Tikus': 15, 'Cap Gajah Lem Tikus': 15,
    'Personal Hand Wash': 15, 'Personal Wet Wipes': 15, 'Home all surface wipes': 15,
    'Home Aerosol': 15, 'Polytex': 20, 'Wetties': 20, 'Personal Liquid Spray': 20,
    'Aluminium Foil': 15, 'Autosol': 20, 'Carrera': 20, 'Clingwrap': 15,
    'Creme': 15, 'Garbage Bag': 15, 'Proclin Aksi Putih': 15, 'Proclin Bleach': 15,
    'Proclin Stain Remover': 15, 'Shampoo HC': 15, 'Shock': 15, 'Susemi': 15,
    'Wonderfuel': 15,
}

# ─────────────────────────────────────────────
# SKU CLASSIFICATION
# ─────────────────────────────────────────────
def classify_sku(row):
    fc = [row['Jun'], row['Jul'], row['Aug'], row['Sep']]
    non_zero = [f > 0 for f in fc]
    if all(non_zero):
        return 'Running'
    if not non_zero[0] and any(non_zero[1:]):
        return 'New Product'
    if non_zero[0] and not all(non_zero):
        # check if phasing out
        last_nonzero = max([i for i, v in enumerate(non_zero) if v], default=-1)
        if last_nonzero < 3:
            return 'Phased Out'
    return 'Running'

df_fc['SKU_Type'] = df_fc.apply(classify_sku, axis=1)

# ─────────────────────────────────────────────
# PRODUCTION PLANNING
# For each month: Production = max(0, Forecast + Target_Closing - Opening)
# where Target_Closing = Forecast_next_month * target_days / days_in_next_month
# if last month, use current month forecast as proxy
# ─────────────────────────────────────────────
month_cols = ['Jun','Jul','Aug','Sep']
month_keys = ['Jun-25','Jul-25','Aug-25','Sep-25']

# Per-checklist production caps: max = avg monthly forecast × 1.20
# (or stated average × 1.20). Ceiling only — no floor.
def _build_prod_caps(df):
    caps = {}
    for cat in df['Category'].unique():
        avg = df[df['Category']==cat][['Jun','Jul','Aug','Sep']].sum().mean()
        caps[cat] = avg * 1.20
    # Override with stated averages from checklist
    caps['Proclin Bleach'] = 14e6 * 1.20
    caps['Pocket']         = 1.8e6 * 1.20
    return caps

def calc_production_plan(df):
    global CATEGORY_PROD_CAPS
    CATEGORY_PROD_CAPS = _build_prod_caps(df)
    results = []
    for _, row in df.iterrows():
        target_days = FG_INV_TARGET.get(row['Category'], 15)
        opening = row['SOH']
        sku_results = {'Category': row['Category'], 'SAP_Code': row['SAP_Code'],
                       'Product_Name': row['Product_Name'], 'SKU_Type': row['SKU_Type']}
        prod_by_month = {}
        opening_by_month = {}
        closing_by_month = {}
        days_cover_by_month = {}

        for i, (mc, mk) in enumerate(zip(month_cols, month_keys)):
            fc = row[mc]
            days = MONTH_DAYS[mk]

            if i < len(month_cols) - 1:
                next_fc = row[month_cols[i+1]]
                next_days = MONTH_DAYS[month_keys[i+1]]
            else:
                next_fc = fc
                next_days = days

            target_closing = (next_fc / next_days) * target_days if next_days > 0 and next_fc > 0 else 0
            # Production needed to cover forecast + target closing
            prod = max(0, fc + target_closing - opening)
            closing = opening + prod - fc
            days_cover = (closing / (fc / days)) if fc > 0 else (closing if closing > 0 else 0)

            opening_by_month[mk] = round(opening, 2)
            prod_by_month[mk] = round(prod, 2)
            closing_by_month[mk] = round(max(0, closing), 2)
            days_cover_by_month[mk] = round(days_cover, 1) if fc > 0 else 0

            opening = closing_by_month[mk]

        sku_results.update({
            'Opening': opening_by_month, 'Production': prod_by_month,
            'Forecast': {mk: row[mc] for mc, mk in zip(month_cols, month_keys)},
            'Closing': closing_by_month, 'Days_Cover': days_cover_by_month
        })
        results.append(sku_results)

    # Post-pass: apply category-level production cap (ceiling only, per checklist).
    # Cap = avg monthly forecast × 1.20.  If category total exceeds cap in any month,
    # scale all SKUs in that category proportionally down for that month,
    # then cascade the revised closing into the next month's opening.
    for mk_idx, mk in enumerate(month_keys):
        cat_groups = {}
        for p in results:
            cat_groups.setdefault(p['Category'], []).append(p)
        for cat, items in cat_groups.items():
            cap = CATEGORY_PROD_CAPS.get(cat)
            if not cap:
                continue
            total = sum(p['Production'][mk] for p in items)
            if total <= cap * 1.001:
                continue
            scale = cap / total
            for p in items:
                new_prod = round(p['Production'][mk] * scale, 2)
                fc_val   = p['Forecast'][mk]
                opening  = p['Opening'][mk]
                new_cl   = round(max(0, opening + new_prod - fc_val), 2)
                days     = MONTH_DAYS[mk]
                new_dc   = round(new_cl / (fc_val / days), 1) if fc_val > 0 else 0
                p['Production'][mk] = new_prod
                p['Closing'][mk]    = new_cl
                p['Days_Cover'][mk] = new_dc
                # Cascade: update next month's opening
                if mk_idx < len(month_keys) - 1:
                    next_mk = month_keys[mk_idx + 1]
                    p['Opening'][next_mk] = new_cl
                    # Recompute next month closing with updated opening
                    next_fc   = p['Forecast'][next_mk]
                    next_prod = p['Production'][next_mk]
                    next_cl   = round(max(0, new_cl + next_prod - next_fc), 2)
                    next_days = MONTH_DAYS[next_mk]
                    next_dc   = round(next_cl / (next_fc / next_days), 1) if next_fc > 0 else 0
                    p['Closing'][next_mk]    = next_cl
                    p['Days_Cover'][next_mk] = next_dc

    return results

print("Calculating production plan...")
prod_plan = calc_production_plan(df_fc)

# ─────────────────────────────────────────────
# MATERIAL PLANNING
# ─────────────────────────────────────────────
print("Calculating material requirements...")

# Build production quantity dict: {sap_code: {month: qty}}
prod_qty = {}
for p in prod_plan:
    prod_qty[p['SAP_Code']] = p['Production']

# Match forecast SKUs to BOM via product name (SAP codes differ between the two files).
# First try exact name match; then fall back to stripping variant suffixes (MTP, NF, etc.)
import re as _re

_bom_name_upper = {
    k.strip().upper().replace('\xa0',' ').replace('\u00a0',' '): v
    for k, v in
    df_bom.drop_duplicates('FG_Name')[['FG_SAP','FG_Name']]
          .set_index('FG_Name')['FG_SAP'].to_dict().items()}

def _bom_match(name):
    n = str(name).strip().upper().replace('\xa0',' ').replace('\u00a0',' ')
    if n in _bom_name_upper:
        return _bom_name_upper[n]
    # Strip variant suffixes and try with/without NF
    base = n
    for suffix in [' MTP RJV', ' MTP', ' NF', ' RJV']:
        base = base.replace(suffix, '').strip()
    # Try base as-is, then base+' NF'
    for cand in [base, base + ' NF', base + ' NF']:
        if cand in _bom_name_upper:
            return _bom_name_upper[cand]
    # Collapse spaces/hyphens entirely and compare
    def _norm(s): return _re.sub(r'[\s\-]+', '', s)
    n_norm = _norm(n)
    for k, v in _bom_name_upper.items():
        if _norm(k) == n_norm:
            return v
    return None

df_fc['_BOM_SAP'] = df_fc['Product_Name'].apply(_bom_match)
matched = df_fc['_BOM_SAP'].notna().sum()
print(f"  Forecast SKUs matched to BOM: {matched}/{len(df_fc)}")

# Build mapping: BOM_SAP -> [forecast SAP codes]
_bom_to_fc = {}
for _, row in df_fc[df_fc['_BOM_SAP'].notna()].iterrows():
    _bom_to_fc.setdefault(row['_BOM_SAP'], []).append(row['SAP_Code'])

# Only use BOM rows whose FG_SAP appears in our name-matched set
_matched_bom_saps = set(_bom_to_fc.keys())
df_bom_fc = df_bom[df_bom['FG_SAP'].isin(_matched_bom_saps)].copy()

# Calculate material requirements per month
mat_req = {}
for _, brow in df_bom_fc.iterrows():
    bom_fg  = brow['FG_SAP']
    comp    = brow['Comp_Code']
    qty_per = brow['Qty_per_FG']
    fc_saps = _bom_to_fc.get(bom_fg, [])
    if not fc_saps:
        continue
    if comp not in mat_req:
        mat_req[comp] = {mk: 0 for mk in month_keys}
    for fc_sap in fc_saps:
        if fc_sap not in prod_qty:
            continue
        for mk in month_keys:
            pq = prod_qty[fc_sap].get(mk, 0)
            mat_req[comp][mk] = mat_req[comp].get(mk, 0) + pq * qty_per

# Now calculate order plan
# Get component master info
comp_master = {}
for _, row in df_csoh.iterrows():
    comp_master[row['Comp_Code']] = {
        'desc': row['Comp_Desc'], 'category': row['Category'],
        'soh': row['SOH'], 'target_days': row['Target_Days'], 'uom': row['UoM']
    }

# Get source (dom/import) per component
comp_source = df_bom.drop_duplicates('Comp_Code').set_index('Comp_Code')['Source'].to_dict()
comp_lead_time = df_bom.drop_duplicates('Comp_Code').set_index('Comp_Code')['Lead_Time'].to_dict()
comp_lot_size = df_bom.drop_duplicates('Comp_Code').set_index('Comp_Code')['Lot_Size'].to_dict()
comp_supplier = df_bom.drop_duplicates('Comp_Code').set_index('Comp_Code')['Supplier_Code'].to_dict()
comp_cat_from_bom = df_bom.drop_duplicates('Comp_Code').set_index('Comp_Code')['Comp_Cat'].to_dict()

def ceil_to_lot(qty, lot_size):
    if lot_size <= 0 or lot_size is None:
        return max(0, round(qty))
    return max(0, math.ceil(qty / lot_size) * lot_size) if qty > 0 else 0

print("Calculating order quantities...")

# Build material plan
mat_plan = {}  # comp_code -> list of monthly dicts

all_comps = set(mat_req.keys()) | set(comp_master.keys())
# Only plan for components that have requirements
comps_with_req = [c for c in mat_req.keys() if sum(mat_req[c].values()) > 0]

for comp in comps_with_req:
    source = comp_source.get(comp, 'Domestic')
    lot_size = comp_lot_size.get(comp, 1)
    lot_size = max(1, lot_size) if lot_size else 1
    lead_time = comp_lead_time.get(comp, 30)
    cat = comp_master.get(comp, {}).get('category', comp_cat_from_bom.get(comp, 'Unknown'))
    target_days_item = comp_master.get(comp, {}).get('target_days', 15)
    
    # For Import: 30 days closing target
    if source == 'Import':
        target_days_item = 30
    
    opening = comp_master.get(comp, {}).get('soh', 0)
    months_data = []
    
    for i, mk in enumerate(month_keys):
        req = mat_req.get(comp, {}).get(mk, 0)
        days = MONTH_DAYS[mk]
        
        # Target closing = requirement/days * target_days
        if req > 0 and days > 0:
            target_close = (req / days) * target_days_item
        else:
            target_close = opening  # keep existing if no req
        
        # Receipt needed
        receipt_need = max(0, req + target_close - opening)
        receipt = ceil_to_lot(receipt_need, lot_size)
        closing = opening + receipt - req
        days_cover = (closing / (req / days)) if req > 0 and days > 0 else (closing if closing > 0 else 0)
        
        months_data.append({
            'month': mk, 'opening': round(opening, 4), 'requirement': round(req, 4),
            'receipt': round(receipt, 4), 'closing': round(max(0, closing), 4),
            'days_cover': round(days_cover, 1) if req > 0 else 0,
            'target_days': target_days_item, 'source': source, 'lot_size': lot_size,
            'lead_time': lead_time, 'category': cat
        })
        opening = max(0, closing)
    
    mat_plan[comp] = months_data

print(f"  Material plan computed for {len(mat_plan)} components")

# ─────────────────────────────────────────────
# CREATE WORKBOOK
# ─────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ═══════════════════════════════════════════════
# SHEET 1: FG Production Plan (SKU Level)
# ═══════════════════════════════════════════════
print("Building Sheet 1: FG Production Plan...")
ws1 = wb.create_sheet("FG Production Plan")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "E5"

# Title
ws1.merge_cells("A1:AG1")
t = ws1["A1"]
t.value = "FINISHED GOODS PRODUCTION PLAN — SKU LEVEL"
t.font = font(bold=True, color=C_WHITE, size=14)
t.fill = fill(C_NAVY)
t.alignment = align(h="center")
ws1.row_dimensions[1].height = 30

# Subtitle row
ws1.merge_cells("A2:AG2")
t2 = ws1["A2"]
t2.value = "Period: Jun-25 to Sep-25  |  Values in Units  |  Days Cover = Closing ÷ Daily Forecast"
t2.font = font(italic=True, color=C_NAVY, size=9)
t2.fill = fill(C_LTBLUE)
t2.alignment = align(h="center")
ws1.row_dimensions[2].height = 16

# Month group headers row 3
month_spans = {}
col_start = 6
for mk in month_keys:
    # 5 cols per month: Opening, Production, Forecast, Closing, Days Cover
    ws1.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_start+4)
    c = ws1.cell(row=3, column=col_start, value=mk)
    c.font = font(bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_TEAL)
    c.alignment = align(h="center")
    c.border = thin_border()
    month_spans[mk] = col_start
    col_start += 5

# Also add last col for SKU Type
ws1.cell(row=3, column=col_start, value="SKU Type").fill = fill(C_NAVY)
ws1.cell(row=3, column=col_start).font = font(bold=True, color=C_WHITE, size=9)
ws1.cell(row=3, column=col_start).alignment = align(h="center")
ws1.cell(row=3, column=col_start).border = thin_border()

# Sub-headers row 4
sub_hdrs = ['#','Category','SAP Code','Product Name','Period']
for mk in month_keys:
    sub_hdrs += ['Opening','Production','Forecast','Closing','Days Cover']
sub_hdrs += ['SKU Type']
apply_subheader(ws1, 4, sub_hdrs, bg=C_BLUE)

# Header row 3 for first 5 cols
for ci in range(1, 6):
    ws1.merge_cells(start_row=3, start_column=ci, end_row=3, end_column=ci)
    c = ws1.cell(row=3, column=ci)
    c.fill = fill(C_NAVY)
    c.border = thin_border()

# Column widths
ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 38
ws1.column_dimensions['E'].width = 12
for col_idx in range(6, col_start+1):
    ws1.column_dimensions[get_column_letter(col_idx)].width = 13

# Data rows
row = 5
prev_cat = None
row_num = 1

NUM_FMT = "#,##0"
DAY_FMT = "#,##0.0"

for p in prod_plan:
    cat = p['Category']
    is_new_cat = cat != prev_cat

    # Category separator row
    if is_new_cat:
        total_cols = 5 + 5*4 + 1
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        cat_cell = ws1.cell(row=row, column=1, value=f"  {cat}")
        cat_cell.font = font(bold=True, color=C_WHITE, size=9)
        cat_cell.fill = fill(C_NAVY)
        cat_cell.alignment = align(h="left")
        ws1.row_dimensions[row].height = 14
        row += 1
        prev_cat = cat

    bg_row = C_WHITE if row_num % 2 == 0 else C_GRAY1

    # Row no, cat, SAP, name, period (empty - we span months)
    write_cell(ws1, row, 1, row_num, bg=bg_row, h="center")
    write_cell(ws1, row, 2, cat, bg=bg_row, h="left", sz=8)
    write_cell(ws1, row, 3, p['SAP_Code'], bg=bg_row, h="center")
    write_cell(ws1, row, 4, p['Product_Name'], bg=bg_row, h="left", sz=8)
    write_cell(ws1, row, 5, "Units", bg=bg_row, h="center", sz=8, italic=True)

    col = 6
    for mk in month_keys:
        op = p['Opening'][mk]
        pr = p['Production'][mk]
        fc = p['Forecast'][mk]
        cl = p['Closing'][mk]
        dc = p['Days_Cover'][mk]

        # Color-code days cover
        dc_bg = C_LTGREEN if dc >= 7 else (C_LTAMBER if dc >= 3 else C_LTRED)

        write_cell(ws1, row, col,   op, bg=bg_row, num_format=NUM_FMT); col+=1
        write_cell(ws1, row, col,   pr, bg=bg_row, num_format=NUM_FMT,
                   bold=(pr > 0), color=C_TEAL if pr > 0 else C_BLACK); col+=1
        write_cell(ws1, row, col,   fc, bg=bg_row, num_format=NUM_FMT); col+=1
        write_cell(ws1, row, col,   cl, bg=bg_row, num_format=NUM_FMT); col+=1
        write_cell(ws1, row, col,   dc, bg=dc_bg, num_format=DAY_FMT, h="center"); col+=1

    # SKU Type
    sku_type = p['SKU_Type']
    st_bg = {'Running': C_LTGREEN, 'New Product': C_LTBLUE, 'Phased Out': C_LTRED}.get(sku_type, C_WHITE)
    st_color = {'Running': C_TEAL, 'New Product': C_BLUE, 'Phased Out': C_RED}.get(sku_type, C_BLACK)
    write_cell(ws1, row, col, sku_type, bg=st_bg, color=st_color, bold=True, h="center", sz=8)

    row += 1
    row_num += 1

print(f"  FG Plan rows written: {row_num-1}")

# ═══════════════════════════════════════════════
# SHEET 2: FG Summary (Category Level)
# ═══════════════════════════════════════════════
print("Building Sheet 2: FG Category Summary...")
ws2 = wb.create_sheet("FG Category Summary")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "B5"

ws2.merge_cells("A1:V1")
t = ws2["A1"]
t.value = "FINISHED GOODS PRODUCTION PLAN — CATEGORY SUMMARY"
t.font = font(bold=True, color=C_WHITE, size=14)
t.fill = fill(C_NAVY)
t.alignment = align(h="center")
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:V2")
t2 = ws2["A2"]
t2.value = "Aggregated by Product Category  |  Period: Jun-25 to Sep-25"
t2.font = font(italic=True, color=C_NAVY, size=9)
t2.fill = fill(C_LTBLUE)
t2.alignment = align(h="center")

# Month headers row 3
col_start = 2
for mk in month_keys:
    ws2.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_start+4)
    c = ws2.cell(row=3, column=col_start, value=mk)
    c.font = font(bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_TEAL)
    c.alignment = align(h="center")
    c.border = thin_border()
    col_start += 5

ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=1)
ws2.cell(row=3, column=1).fill = fill(C_NAVY)
ws2.cell(row=3, column=1).border = thin_border()

sub2 = ['Category']
for mk in month_keys:
    sub2 += ['Opening','Production','Forecast','Closing','Days Cover']
apply_subheader(ws2, 4, sub2, bg=C_BLUE)

ws2.column_dimensions['A'].width = 30
for ci in range(2, 22):
    ws2.column_dimensions[get_column_letter(ci)].width = 13

# Aggregate by category
cat_summary = {}
for p in prod_plan:
    cat = p['Category']
    if cat not in cat_summary:
        cat_summary[cat] = {mk: {'op':0,'pr':0,'fc':0,'cl':0} for mk in month_keys}
    for mk in month_keys:
        cat_summary[cat][mk]['op'] += p['Opening'][mk]
        cat_summary[cat][mk]['pr'] += p['Production'][mk]
        cat_summary[cat][mk]['fc'] += p['Forecast'][mk]
        cat_summary[cat][mk]['cl'] += p['Closing'][mk]

row = 5
grand = {mk: {'op':0,'pr':0,'fc':0,'cl':0} for mk in month_keys}

for i, (cat, data) in enumerate(sorted(cat_summary.items())):
    bg_row = C_WHITE if i % 2 == 0 else C_GRAY1
    write_cell(ws2, row, 1, cat, bg=bg_row, bold=True, h="left", sz=9)
    col = 2
    for mk in month_keys:
        op = data[mk]['op']; pr = data[mk]['pr']
        fc = data[mk]['fc']; cl = data[mk]['cl']
        days = MONTH_DAYS[mk]
        dc = (cl / (fc/days)) if fc > 0 else 0
        dc_bg = C_LTGREEN if dc >= 7 else (C_LTAMBER if dc >= 3 else C_LTRED)
        write_cell(ws2, row, col, op, bg=bg_row, num_format=NUM_FMT); col+=1
        write_cell(ws2, row, col, pr, bg=bg_row, num_format=NUM_FMT, bold=True, color=C_TEAL); col+=1
        write_cell(ws2, row, col, fc, bg=bg_row, num_format=NUM_FMT); col+=1
        write_cell(ws2, row, col, cl, bg=bg_row, num_format=NUM_FMT); col+=1
        write_cell(ws2, row, col, round(dc,1), bg=dc_bg, num_format=DAY_FMT, h="center"); col+=1
        grand[mk]['op'] += op; grand[mk]['pr'] += pr
        grand[mk]['fc'] += fc; grand[mk]['cl'] += cl
    row += 1

# Grand total
write_cell(ws2, row, 1, "GRAND TOTAL", bg=C_NAVY, bold=True, color=C_WHITE, h="left", sz=10)
col = 2
for mk in month_keys:
    op = grand[mk]['op']; pr = grand[mk]['pr']
    fc = grand[mk]['fc']; cl = grand[mk]['cl']
    days = MONTH_DAYS[mk]
    dc = (cl / (fc/days)) if fc > 0 else 0
    dc_bg = C_LTGREEN if dc >= 7 else (C_LTAMBER if dc >= 3 else C_LTRED)
    write_cell(ws2, row, col, op, bg=C_NAVY, bold=True, color=C_WHITE, num_format=NUM_FMT); col+=1
    write_cell(ws2, row, col, pr, bg=C_NAVY, bold=True, color=C_WHITE, num_format=NUM_FMT); col+=1
    write_cell(ws2, row, col, fc, bg=C_NAVY, bold=True, color=C_WHITE, num_format=NUM_FMT); col+=1
    write_cell(ws2, row, col, cl, bg=C_NAVY, bold=True, color=C_WHITE, num_format=NUM_FMT); col+=1
    write_cell(ws2, row, col, round(dc,1), bg=C_TEAL, bold=True, color=C_WHITE, num_format=DAY_FMT, h="center"); col+=1

# ═══════════════════════════════════════════════
# SHEET 3: FG Executive Summary
# ═══════════════════════════════════════════════
print("Building Sheet 3: Executive Summary...")
ws3 = wb.create_sheet("Executive Summary")
ws3.sheet_view.showGridLines = False

ws3.column_dimensions['A'].width = 3
ws3.column_dimensions['B'].width = 28
for ci in range(3, 20):
    ws3.column_dimensions[get_column_letter(ci)].width = 14

# Title
ws3.merge_cells("B1:R1")
t = ws3["B1"]
t.value = "INTEGRATED SUPPLY & FORECAST PLAN — EXECUTIVE SUMMARY"
t.font = font(bold=True, color=C_WHITE, size=14)
t.fill = fill(C_NAVY)
t.alignment = align(h="center")
ws3.row_dimensions[1].height = 35

ws3.merge_cells("B2:R2")
t2 = ws3["B2"]
t2.value = "Plan Period: Jun-2025 to Sep-2025  |  Auto-generated from SAP Forecast & BOM Data"
t2.font = font(italic=True, color=C_NAVY, size=9)
t2.fill = fill(C_LTBLUE)
t2.alignment = align(h="center")

# Section 1: SKU Classification summary
ws3.merge_cells("B4:R4")
s1 = ws3["B4"]
s1.value = "▶  SECTION 1: SKU PORTFOLIO OVERVIEW"
s1.font = font(bold=True, color=C_WHITE, size=10)
s1.fill = fill(C_TEAL)
s1.alignment = align(h="left")

sku_counts = df_fc['SKU_Type'].value_counts()
hdr5 = ['SKU Type', 'Count', 'Notes']
for ci, h in enumerate(hdr5, start=2):
    c = ws3.cell(row=5, column=ci, value=h)
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_BLUE)
    c.alignment = align(h="center"); c.border = thin_border()

types_info = [
    ('Running', sku_counts.get('Running', 0), 'Forecast available all 4 months — steady production required'),
    ('New Product', sku_counts.get('New Product', 0), 'Launch in later months — ramp-up production needed'),
    ('Phased Out', sku_counts.get('Phased Out', 0), 'Forecast in early months only — declining production'),
]
for ri, (t, cnt, note) in enumerate(types_info, start=6):
    bg = C_LTGREEN if t=='Running' else (C_LTBLUE if t=='New Product' else C_LTRED)
    write_cell(ws3, ri, 2, t, bg=bg, bold=True, h="left")
    write_cell(ws3, ri, 3, cnt, bg=bg, bold=True, h="center")
    write_cell(ws3, ri, 4, note, bg=bg, h="left", sz=8)
    ws3.merge_cells(start_row=ri, start_column=4, end_row=ri, end_column=18)

# Section 2: Monthly Forecast vs Production
ws3.merge_cells("B10:R10")
s2 = ws3["B10"]
s2.value = "▶  SECTION 2: MONTHLY FORECAST vs PRODUCTION (All Categories)"
s2.font = font(bold=True, color=C_WHITE, size=10)
s2.fill = fill(C_TEAL)
s2.alignment = align(h="left")

hdr11 = ['Metric'] + MONTHS + ['Total']
for ci, h in enumerate(hdr11, start=2):
    c = ws3.cell(row=11, column=ci, value=h)
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_BLUE)
    c.alignment = align(h="center"); c.border = thin_border()

totals_by_metric = {
    'Total Forecast': {mk: sum(p['Forecast'][mk] for p in prod_plan) for mk in month_keys},
    'Total Production': {mk: sum(p['Production'][mk] for p in prod_plan) for mk in month_keys},
    'FG Opening Inv.': {mk: sum(p['Opening'][mk] for p in prod_plan) for mk in month_keys},
    'FG Closing Inv.': {mk: sum(p['Closing'][mk] for p in prod_plan) for mk in month_keys},
    'Overall Days Cover': None,  # compute separately
}

metrics_rows = [
    ('Total Forecast', C_WHITE, C_BLACK),
    ('Total Production', C_LTGREEN, C_TEAL),
    ('FG Opening Inv.', C_GRAY1, C_BLACK),
    ('FG Closing Inv.', C_LTBLUE, C_BLUE),
]
for ri, (metric, bg, fg) in enumerate(metrics_rows, start=12):
    vals = totals_by_metric[metric]
    write_cell(ws3, ri, 2, metric, bg=bg, bold=True, color=fg, h="left")
    total_sum = 0
    for ci, mk in enumerate(month_keys, start=3):
        v = vals[mk]
        total_sum += v
        write_cell(ws3, ri, ci, round(v), bg=bg, color=fg, num_format=NUM_FMT)
    write_cell(ws3, ri, 7, round(total_sum), bg=bg, bold=True, color=fg, num_format=NUM_FMT)

# Days cover row
ri = 16
write_cell(ws3, ri, 2, 'Avg Days Cover', bg=C_LTAMBER, bold=True, h="left")
for ci, mk in enumerate(month_keys, start=3):
    fc_total = totals_by_metric['Total Forecast'][mk]
    cl_total = totals_by_metric['FG Closing Inv.'][mk]
    days = MONTH_DAYS[mk]
    dc = (cl_total / (fc_total/days)) if fc_total > 0 else 0
    dc_bg = C_LTGREEN if dc >= 7 else (C_LTAMBER if dc >= 3 else C_LTRED)
    write_cell(ws3, ri, ci, round(dc, 1), bg=dc_bg, num_format=DAY_FMT, h="center")
ws3.merge_cells(start_row=ri, start_column=7, end_row=ri, end_column=7)

# Section 3: Top categories by volume
ws3.merge_cells("B19:R19")
s3 = ws3["B19"]
s3.value = "▶  SECTION 3: TOP CATEGORIES BY FORECAST VOLUME"
s3.font = font(bold=True, color=C_WHITE, size=10)
s3.fill = fill(C_TEAL)
s3.alignment = align(h="left")

hdr20 = ['Category', 'Jun-25', 'Jul-25', 'Aug-25', 'Sep-25', '4-Month Total', '% of Total']
for ci, h in enumerate(hdr20, start=2):
    c = ws3.cell(row=20, column=ci, value=h)
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_BLUE)
    c.alignment = align(h="center"); c.border = thin_border()

cat_totals_list = sorted(
    [(cat, sum(data[mk]['fc'] for mk in month_keys)) 
     for cat, data in cat_summary.items()],
    key=lambda x: -x[1])

overall_total = sum(x[1] for x in cat_totals_list)
for ri, (cat, tot) in enumerate(cat_totals_list[:15], start=21):
    bg = C_WHITE if ri % 2 == 0 else C_GRAY1
    write_cell(ws3, ri, 2, cat, bg=bg, h="left", bold=True, sz=8)
    for ci, mk in enumerate(month_keys, start=3):
        write_cell(ws3, ri, ci, round(cat_summary[cat][mk]['fc']), bg=bg, num_format=NUM_FMT)
    write_cell(ws3, ri, 7, round(tot), bg=bg, bold=True, num_format=NUM_FMT)
    pct = tot / overall_total * 100 if overall_total > 0 else 0
    write_cell(ws3, ri, 8, round(pct, 1), bg=bg, num_format="0.0\"%\"", h="center")

# Section 4: Key Alerts
alert_row = 38
ws3.merge_cells(f"B{alert_row}:R{alert_row}")
s4 = ws3[f"B{alert_row}"]
s4.value = "▶  SECTION 4: CRITICAL ALERTS & PLANNING FLAGS"
s4.font = font(bold=True, color=C_WHITE, size=10)
s4.fill = fill(C_RED)
s4.alignment = align(h="left")

alert_hdr = ['Alert Type', 'SAP Code', 'Product', 'Category', 'Month', 'Days Cover', 'Action Required']
for ci, h in enumerate(alert_hdr, start=2):
    c = ws3.cell(row=alert_row+1, column=ci, value=h)
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_RED)
    c.alignment = align(h="center"); c.border = thin_border()

alert_row_data = alert_row + 2
alert_count = 0
for p in prod_plan:
    for mk in month_keys:
        dc = p['Days_Cover'][mk]
        fc = p['Forecast'][mk]
        if fc > 0 and dc < 5 and alert_count < 30:
            alert_type = "⚠ LOW STOCK" if dc < 3 else "! BELOW TARGET"
            alert_bg = C_LTRED if dc < 3 else C_LTAMBER
            write_cell(ws3, alert_row_data, 2, alert_type, bg=alert_bg, bold=True, color=C_RED, sz=8, h="center")
            write_cell(ws3, alert_row_data, 3, p['SAP_Code'], bg=alert_bg, h="center", sz=8)
            write_cell(ws3, alert_row_data, 4, p['Product_Name'], bg=alert_bg, h="left", sz=7)
            write_cell(ws3, alert_row_data, 5, p['Category'], bg=alert_bg, h="left", sz=8)
            write_cell(ws3, alert_row_data, 6, mk, bg=alert_bg, h="center", sz=8)
            write_cell(ws3, alert_row_data, 7, round(dc,1), bg=alert_bg, h="center", num_format=DAY_FMT)
            action = "Increase production" if dc < 3 else "Monitor closely"
            write_cell(ws3, alert_row_data, 8, action, bg=alert_bg, h="left", sz=8)
            alert_row_data += 1
            alert_count += 1

if alert_count == 0:
    ws3.merge_cells(f"B{alert_row_data}:R{alert_row_data}")
    c = ws3.cell(row=alert_row_data, column=2, value="✓ No critical stock alerts — all SKUs above minimum threshold")
    c.font = font(color=C_TEAL, bold=True, size=9)
    c.fill = fill(C_LTGREEN)
    c.alignment = align(h="center")

# ═══════════════════════════════════════════════
# SHEET 4: Material Plan (Component Level)
# ═══════════════════════════════════════════════
print("Building Sheet 4: Material Plan Component Level...")
ws4 = wb.create_sheet("Material Plan - Component")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "E5"

ws4.merge_cells("A1:AH1")
t = ws4["A1"]
t.value = "MATERIAL REQUIREMENT PLAN — COMPONENT LEVEL"
t.font = font(bold=True, color=C_WHITE, size=14)
t.fill = fill(C_NAVY)
t.alignment = align(h="center")
ws4.row_dimensions[1].height = 30

ws4.merge_cells("A2:AH2")
t2 = ws4["A2"]
t2.value = "Values in component UoM  |  Receipt = Order Quantity (rounded to lot size)  |  Import target: 30 days  |  Domestic: Category target days"
t2.font = font(italic=True, color=C_NAVY, size=9)
t2.fill = fill(C_LTBLUE)
t2.alignment = align(h="center")

# Month group headers
col_start = 8
for mk in month_keys:
    ws4.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_start+4)
    c = ws4.cell(row=3, column=col_start, value=mk)
    c.font = font(bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_PURPLE)
    c.alignment = align(h="center"); c.border = thin_border()
    col_start += 5

# Fix cols 1-7
for ci in range(1, 8):
    ws4.cell(row=3, column=ci).fill = fill(C_NAVY)
    ws4.cell(row=3, column=ci).border = thin_border()

sub4 = ['#','Comp Code','Description','Category','Source','Lead Time','Lot Size']
for mk in month_keys:
    sub4 += ['Opening','Requirement','Receipt','Closing','Days Cover']
apply_subheader(ws4, 4, sub4, bg=C_PURPLE)

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 35
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 10
ws4.column_dimensions['F'].width = 11
ws4.column_dimensions['G'].width = 11
for ci in range(8, 29):
    ws4.column_dimensions[get_column_letter(ci)].width = 13

row = 5
comp_row_num = 1
prev_cat = None

# Sort by category then component
sorted_comps = sorted(mat_plan.items(), key=lambda x: (
    comp_master.get(x[0], {}).get('category', ''), x[0]))

for comp, months_data in sorted_comps:
    cat = months_data[0]['category']
    source = months_data[0]['source']
    lead_time = months_data[0]['lead_time']
    lot_size = months_data[0]['lot_size']
    desc = comp_master.get(comp, {}).get('desc', 'Unknown')

    is_new_cat = cat != prev_cat
    if is_new_cat:
        total_cols = 7 + 5*4
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        cc = ws4.cell(row=row, column=1, value=f"  {cat}")
        cc.font = font(bold=True, color=C_WHITE, size=9)
        cc.fill = fill(C_PURPLE)
        cc.alignment = align(h="left")
        ws4.row_dimensions[row].height = 14
        row += 1
        prev_cat = cat

    bg_row = C_WHITE if comp_row_num % 2 == 0 else C_GRAY1
    src_color = C_RED if source == 'Import' else C_TEAL

    write_cell(ws4, row, 1, comp_row_num, bg=bg_row, h="center")
    write_cell(ws4, row, 2, comp, bg=bg_row, h="center", sz=8)
    write_cell(ws4, row, 3, desc, bg=bg_row, h="left", sz=8)
    write_cell(ws4, row, 4, cat, bg=bg_row, h="left", sz=8)
    write_cell(ws4, row, 5, source, bg=C_LTRED if source=='Import' else C_LTGREEN,
               color=src_color, bold=True, h="center", sz=8)
    write_cell(ws4, row, 6, f"{int(lead_time)}d", bg=bg_row, h="center", sz=8)
    write_cell(ws4, row, 7, lot_size, bg=bg_row, h="center", sz=8, num_format=NUM_FMT)

    col = 8
    for md in months_data:
        dc_bg = C_LTGREEN if md['days_cover'] >= 15 else (C_LTAMBER if md['days_cover'] >= 5 else C_LTRED)
        if md['requirement'] == 0:
            dc_bg = C_GRAY1

        write_cell(ws4, row, col,   md['opening'], bg=bg_row, num_format="#,##0.00"); col+=1
        write_cell(ws4, row, col,   md['requirement'], bg=bg_row, num_format="#,##0.00"); col+=1
        write_cell(ws4, row, col,   md['receipt'], bg=bg_row, num_format=NUM_FMT,
                   bold=(md['receipt']>0), color=C_PURPLE if md['receipt']>0 else C_BLACK); col+=1
        write_cell(ws4, row, col,   md['closing'], bg=bg_row, num_format="#,##0.00"); col+=1
        write_cell(ws4, row, col,   md['days_cover'], bg=dc_bg, num_format=DAY_FMT, h="center"); col+=1

    row += 1
    comp_row_num += 1

print(f"  Material plan rows: {comp_row_num-1}")

# ═══════════════════════════════════════════════
# SHEET 5: Material Plan Category Summary
# ═══════════════════════════════════════════════
print("Building Sheet 5: Material Category Summary...")
ws5 = wb.create_sheet("Material Plan - Category")
ws5.sheet_view.showGridLines = False
ws5.freeze_panes = "B5"

ws5.merge_cells("A1:V1")
t = ws5["A1"]
t.value = "MATERIAL REQUIREMENT PLAN — CATEGORY SUMMARY"
t.font = font(bold=True, color=C_WHITE, size=14)
t.fill = fill(C_NAVY)
t.alignment = align(h="center")
ws5.row_dimensions[1].height = 30

ws5.merge_cells("A2:V2")
ws5["A2"].value = "Aggregated by Component Material Group  |  Period: Jun-25 to Sep-25"
ws5["A2"].font = font(italic=True, color=C_NAVY, size=9)
ws5["A2"].fill = fill(C_LTPURPLE)
ws5["A2"].alignment = align(h="center")

col_start = 2
for mk in month_keys:
    ws5.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_start+4)
    c = ws5.cell(row=3, column=col_start, value=mk)
    c.font = font(bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_PURPLE)
    c.alignment = align(h="center"); c.border = thin_border()
    col_start += 5

ws5.cell(row=3, column=1).fill = fill(C_NAVY); ws5.cell(row=3, column=1).border = thin_border()

sub5 = ['Material Group']
for mk in month_keys:
    sub5 += ['Opening','Requirement','Receipt','Closing','Days Cover']
apply_subheader(ws5, 4, sub5, bg=C_PURPLE)

ws5.column_dimensions['A'].width = 28
for ci in range(2, 22):
    ws5.column_dimensions[get_column_letter(ci)].width = 14

# Aggregate
cat_mat = {}
for comp, months_data in mat_plan.items():
    cat = months_data[0]['category']
    if cat not in cat_mat:
        cat_mat[cat] = {mk: {'op':0,'req':0,'rec':0,'cl':0} for mk in month_keys}
    for md in months_data:
        mk = md['month']
        cat_mat[cat][mk]['op'] += md['opening']
        cat_mat[cat][mk]['req'] += md['requirement']
        cat_mat[cat][mk]['rec'] += md['receipt']
        cat_mat[cat][mk]['cl'] += md['closing']

row = 5
grand5 = {mk: {'op':0,'req':0,'rec':0,'cl':0} for mk in month_keys}

for i, (cat, data) in enumerate(sorted(cat_mat.items())):
    bg_row = C_WHITE if i % 2 == 0 else C_GRAY1
    write_cell(ws5, row, 1, cat, bg=bg_row, bold=True, h="left", sz=9)
    col = 2
    for mk in month_keys:
        op = data[mk]['op']; req = data[mk]['req']
        rec = data[mk]['rec']; cl = data[mk]['cl']
        days = MONTH_DAYS[mk]
        dc = (cl / (req/days)) if req > 0 else 0
        dc_bg = C_LTGREEN if dc >= 15 else (C_LTAMBER if dc >= 5 else C_LTRED)
        write_cell(ws5, row, col, round(op), bg=bg_row, num_format=NUM_FMT); col+=1
        write_cell(ws5, row, col, round(req), bg=bg_row, num_format=NUM_FMT); col+=1
        write_cell(ws5, row, col, round(rec), bg=bg_row, bold=True, color=C_PURPLE, num_format=NUM_FMT); col+=1
        write_cell(ws5, row, col, round(cl), bg=bg_row, num_format=NUM_FMT); col+=1
        write_cell(ws5, row, col, round(dc,1), bg=dc_bg, num_format=DAY_FMT, h="center"); col+=1
        grand5[mk]['op']+=op; grand5[mk]['req']+=req; grand5[mk]['rec']+=rec; grand5[mk]['cl']+=cl
    row += 1

write_cell(ws5, row, 1, "GRAND TOTAL", bg=C_NAVY, bold=True, color=C_WHITE, h="left", sz=10)
col = 2
for mk in month_keys:
    op=grand5[mk]['op']; req=grand5[mk]['req']; rec=grand5[mk]['rec']; cl=grand5[mk]['cl']
    days=MONTH_DAYS[mk]; dc=(cl/(req/days)) if req>0 else 0
    write_cell(ws5, row, col, round(op), bg=C_NAVY, bold=True, color=C_WHITE, num_format=NUM_FMT); col+=1
    write_cell(ws5, row, col, round(req), bg=C_NAVY, bold=True, color=C_WHITE, num_format=NUM_FMT); col+=1
    write_cell(ws5, row, col, round(rec), bg=C_NAVY, bold=True, color=C_WHITE, num_format=NUM_FMT); col+=1
    write_cell(ws5, row, col, round(cl), bg=C_NAVY, bold=True, color=C_WHITE, num_format=NUM_FMT); col+=1
    write_cell(ws5, row, col, round(dc,1), bg=C_TEAL, bold=True, color=C_WHITE, num_format=DAY_FMT, h="center"); col+=1

# ═══════════════════════════════════════════════
# SHEET 6: Domestic Material Tracker
# ═══════════════════════════════════════════════
print("Building Sheet 6: Domestic Material Tracker...")
ws6 = wb.create_sheet("Tracker - Domestic")
ws6.sheet_view.showGridLines = False
ws6.freeze_panes = "F4"

ws6.merge_cells("A1:R1")
t = ws6["A1"]
t.value = "MATERIAL PROCUREMENT TRACKER — DOMESTIC ITEMS"
t.font = font(bold=True, color=C_WHITE, size=13)
t.fill = fill(C_TEAL)
t.alignment = align(h="center")
ws6.row_dimensions[1].height = 28

ws6.merge_cells("A2:R2")
ws6["A2"].value = "Review with procurement team to ensure timely ordering. Order Date = Receipt Month − Lead Time (days)"
ws6["A2"].font = font(italic=True, color=C_TEAL, size=9)
ws6["A2"].fill = fill(C_LTGREEN)
ws6["A2"].alignment = align(h="center")

tracker_hdrs = ['#','Comp Code','Description','Category','Supplier','Lead\nTime',
                'Lot\nSize','Jun-25\nOrder','Jun-25\nOrder By','Jul-25\nOrder',
                'Jul-25\nOrder By','Aug-25\nOrder','Aug-25\nOrder By','Sep-25\nOrder',
                'Sep-25\nOrder By','Status','ABC\nClass','Notes']
apply_subheader(ws6, 3, tracker_hdrs, bg=C_TEAL)

ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 12
ws6.column_dimensions['C'].width = 32
ws6.column_dimensions['D'].width = 20
ws6.column_dimensions['E'].width = 28
ws6.column_dimensions['F'].width = 8
ws6.column_dimensions['G'].width = 10
for ci in range(8, 16):
    ws6.column_dimensions[get_column_letter(ci)].width = 13
ws6.column_dimensions['P'].width = 14
ws6.column_dimensions['Q'].width = 9
ws6.column_dimensions['R'].width = 22

# Month start dates for order-by calculation
import datetime
month_starts = {
    'Jun-25': datetime.date(2025,6,1), 'Jul-25': datetime.date(2025,7,1),
    'Aug-25': datetime.date(2025,8,1), 'Sep-25': datetime.date(2025,9,1)
}

dom_row = 4
dom_num = 1

dom_comps = [(c, m) for c, m in sorted_comps if m[0]['source']=='Domestic']

abc_map = df_prices.set_index('Comp_Code')['ABC_Class'].to_dict()

for comp, months_data in dom_comps:
    desc = comp_master.get(comp, {}).get('desc', 'Unknown')
    cat = months_data[0]['category']
    lead_time = int(months_data[0]['lead_time'])
    lot_size = months_data[0]['lot_size']
    supplier = str(comp_supplier.get(comp, 'TBD'))
    abc_class = abc_map.get(comp, '-')

    # Determine if any orders needed
    has_orders = any(md['receipt'] > 0 for md in months_data)
    bg_row = C_WHITE if dom_num % 2 == 0 else C_GRAY1

    write_cell(ws6, dom_row, 1, dom_num, bg=bg_row, h="center")
    write_cell(ws6, dom_row, 2, comp, bg=bg_row, h="center", sz=8)
    write_cell(ws6, dom_row, 3, desc, bg=bg_row, h="left", sz=8)
    write_cell(ws6, dom_row, 4, cat, bg=bg_row, h="left", sz=8)
    write_cell(ws6, dom_row, 5, supplier, bg=bg_row, h="left", sz=8)
    write_cell(ws6, dom_row, 6, f"{lead_time}d", bg=bg_row, h="center", sz=8)
    write_cell(ws6, dom_row, 7, lot_size, bg=bg_row, h="center", sz=8, num_format=NUM_FMT)

    col = 8
    for md in months_data:
        mk = md['month']
        order_qty = md['receipt']
        # Order-by date = start of month - lead time
        ms = month_starts[mk]
        order_by = ms - datetime.timedelta(days=lead_time)
        order_by_str = order_by.strftime("%d-%b-%y")

        order_bg = C_LTGREEN if order_qty > 0 else C_GRAY1
        order_color = C_TEAL if order_qty > 0 else C_GRAY3
        write_cell(ws6, dom_row, col, order_qty if order_qty > 0 else 0, bg=order_bg,
                   color=order_color, num_format=NUM_FMT, bold=(order_qty>0)); col+=1
        write_cell(ws6, dom_row, col, order_by_str if order_qty > 0 else "-",
                   bg=order_bg, color=order_color, h="center", sz=8); col+=1

    # Status
    total_req = sum(md['requirement'] for md in months_data)
    status = "Needs Order" if has_orders else ("No Req" if total_req == 0 else "OK - SOH Covers")
    st_bg = C_LTAMBER if has_orders else (C_GRAY1 if total_req==0 else C_LTGREEN)
    st_color = C_AMBER if has_orders else (C_GRAY3 if total_req==0 else C_TEAL)
    write_cell(ws6, dom_row, 16, status, bg=st_bg, color=st_color, bold=True, h="center", sz=8)

    abc_bg = {'A': C_LTRED, 'B': C_LTAMBER, 'C': C_LTGREEN}.get(str(abc_class), C_GRAY1)
    abc_color = {'A': C_RED, 'B': C_AMBER, 'C': C_TEAL}.get(str(abc_class), C_BLACK)
    write_cell(ws6, dom_row, 17, abc_class, bg=abc_bg, color=abc_color, bold=True, h="center")
    write_cell(ws6, dom_row, 18, "", bg=bg_row, h="left")

    dom_row += 1
    dom_num += 1

print(f"  Domestic tracker rows: {dom_num-1}")

# ═══════════════════════════════════════════════
# SHEET 7: Import Material Tracker
# ═══════════════════════════════════════════════
print("Building Sheet 7: Import Material Tracker...")
ws7 = wb.create_sheet("Tracker - Import")
ws7.sheet_view.showGridLines = False
ws7.freeze_panes = "F4"

ws7.merge_cells("A1:R1")
t = ws7["A1"]
t.value = "MATERIAL PROCUREMENT TRACKER — IMPORT ITEMS (30-DAY CLOSING TARGET)"
t.font = font(bold=True, color=C_WHITE, size=13)
t.fill = fill(C_RED)
t.alignment = align(h="center")
ws7.row_dimensions[1].height = 28

ws7.merge_cells("A2:R2")
ws7["A2"].value = "Import items have higher lead times — order well in advance. Closing inventory target: 30 days."
ws7["A2"].font = font(italic=True, color=C_RED, size=9)
ws7["A2"].fill = fill(C_LTRED)
ws7["A2"].alignment = align(h="center")

apply_subheader(ws7, 3, tracker_hdrs, bg=C_RED)
for ci_ltr in ['A','B','C','D','E','F','G','P','Q','R']:
    ws7.column_dimensions[ci_ltr].width = ws6.column_dimensions[ci_ltr].width
for ci in range(8, 16):
    ws7.column_dimensions[get_column_letter(ci)].width = 13

imp_comps = [(c, m) for c, m in sorted_comps if m[0]['source']=='Import']

imp_row = 4
imp_num = 1
for comp, months_data in imp_comps:
    desc = comp_master.get(comp, {}).get('desc', 'Unknown')
    cat = months_data[0]['category']
    lead_time = int(months_data[0]['lead_time'])
    lot_size = months_data[0]['lot_size']
    supplier = str(comp_supplier.get(comp, 'TBD'))
    abc_class = abc_map.get(comp, '-')
    has_orders = any(md['receipt'] > 0 for md in months_data)
    bg_row = C_WHITE if imp_num % 2 == 0 else C_GRAY1

    write_cell(ws7, imp_row, 1, imp_num, bg=bg_row, h="center")
    write_cell(ws7, imp_row, 2, comp, bg=bg_row, h="center", sz=8)
    write_cell(ws7, imp_row, 3, desc, bg=bg_row, h="left", sz=8)
    write_cell(ws7, imp_row, 4, cat, bg=bg_row, h="left", sz=8)
    write_cell(ws7, imp_row, 5, supplier, bg=bg_row, h="left", sz=8)
    write_cell(ws7, imp_row, 6, f"{lead_time}d", bg=C_LTRED, h="center", sz=8, color=C_RED, bold=True)
    write_cell(ws7, imp_row, 7, lot_size, bg=bg_row, h="center", sz=8, num_format=NUM_FMT)

    col = 8
    for md in months_data:
        mk = md['month']
        order_qty = md['receipt']
        ms = month_starts[mk]
        order_by = ms - datetime.timedelta(days=lead_time)
        order_by_str = order_by.strftime("%d-%b-%y")

        order_bg = C_LTRED if order_qty > 0 else C_GRAY1
        order_color = C_RED if order_qty > 0 else C_GRAY3
        write_cell(ws7, imp_row, col, order_qty if order_qty > 0 else 0, bg=order_bg,
                   color=order_color, num_format=NUM_FMT, bold=(order_qty>0)); col+=1
        write_cell(ws7, imp_row, col, order_by_str if order_qty > 0 else "-",
                   bg=order_bg, color=order_color, h="center", sz=8); col+=1

    total_req = sum(md['requirement'] for md in months_data)
    status = "Needs Order" if has_orders else ("No Req" if total_req == 0 else "OK - SOH Covers")
    st_bg = C_LTRED if has_orders else (C_GRAY1 if total_req==0 else C_LTGREEN)
    write_cell(ws7, imp_row, 16, status, bg=st_bg, color=C_RED if has_orders else C_TEAL,
               bold=True, h="center", sz=8)
    abc_bg = {'A': C_LTRED, 'B': C_LTAMBER, 'C': C_LTGREEN}.get(str(abc_class), C_GRAY1)
    abc_color = {'A': C_RED, 'B': C_AMBER, 'C': C_TEAL}.get(str(abc_class), C_BLACK)
    write_cell(ws7, imp_row, 17, abc_class, bg=abc_bg, color=abc_color, bold=True, h="center")
    write_cell(ws7, imp_row, 18, "", bg=bg_row)

    imp_row += 1
    imp_num += 1

print(f"  Import tracker rows: {imp_num-1}")

# ═══════════════════════════════════════════════
# SHEET 8: Supplier PO Schedule
# ═══════════════════════════════════════════════
print("Building Sheet 8: Supplier Schedule...")
ws8 = wb.create_sheet("Supplier Schedule")
ws8.sheet_view.showGridLines = False
ws8.freeze_panes = "D4"

ws8.merge_cells("A1:L1")
t = ws8["A1"]
t.value = "SUPPLIER-WISE MATERIAL REQUIREMENT SCHEDULE"
t.font = font(bold=True, color=C_WHITE, size=13)
t.fill = fill(C_NAVY)
t.alignment = align(h="center")
ws8.row_dimensions[1].height = 28

ws8.merge_cells("A2:L2")
ws8["A2"].value = "Monthly order quantities by supplier. Send to suppliers at beginning of each month."
ws8["A2"].font = font(italic=True, color=C_NAVY, size=9)
ws8["A2"].fill = fill(C_LTBLUE)
ws8["A2"].alignment = align(h="center")

sup_hdrs = ['Supplier Code','Supplier Name','Comp Code','Description','Category','Source',
            'Unit','Jun-25','Jul-25','Aug-25','Sep-25','Total (4M)']
apply_subheader(ws8, 3, sup_hdrs, bg=C_NAVY)

ws8.column_dimensions['A'].width = 14
ws8.column_dimensions['B'].width = 30
ws8.column_dimensions['C'].width = 12
ws8.column_dimensions['D'].width = 35
ws8.column_dimensions['E'].width = 20
ws8.column_dimensions['F'].width = 10
ws8.column_dimensions['G'].width = 8
for ci in range(8, 13):
    ws8.column_dimensions[get_column_letter(ci)].width = 14

# Build supplier-wise data
sup_data = []
for comp, months_data in sorted_comps:
    supplier = str(comp_supplier.get(comp, 'UNKNOWN'))
    desc = comp_master.get(comp, {}).get('desc', 'Unknown')
    cat = months_data[0]['category']
    source = months_data[0]['source']
    uom = comp_master.get(comp, {}).get('uom', '-')
    orders = {md['month']: md['receipt'] for md in months_data}
    if any(v > 0 for v in orders.values()):
        sup_data.append({
            'supplier': supplier, 'desc': desc, 'comp': comp, 'cat': cat,
            'source': source, 'uom': uom, 'orders': orders
        })

sup_data_sorted = sorted(sup_data, key=lambda x: (x['supplier'], x['source'], x['cat']))

row = 4
prev_sup = None
sup_num = 1
for item in sup_data_sorted:
    sup = item['supplier']
    source = item['source']

    is_new_sup = sup != prev_sup
    if is_new_sup:
        # Supplier header
        ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        sup_name = 'Unknown'
        if sup in df_vendor['Supplier Code'].values:
            sup_name = df_vendor[df_vendor['Supplier Code']==sup]['Supplier Name'].iloc[0]
        sc = ws8.cell(row=row, column=1, value=f"  Supplier: {sup} — {sup_name}")
        sc.font = font(bold=True, color=C_WHITE, size=9)
        sc.fill = fill(C_TEAL if source=='Domestic' else C_RED)
        sc.alignment = align(h="left")
        ws8.row_dimensions[row].height = 14
        row += 1
        prev_sup = sup

    bg_row = C_WHITE if sup_num % 2 == 0 else C_GRAY1
    write_cell(ws8, row, 1, sup, bg=bg_row, h="center", sz=8)
    write_cell(ws8, row, 2, "", bg=bg_row)  # supplier name col - empty for data rows
    write_cell(ws8, row, 3, item['comp'], bg=bg_row, h="center", sz=8)
    write_cell(ws8, row, 4, item['desc'], bg=bg_row, h="left", sz=8)
    write_cell(ws8, row, 5, item['cat'], bg=bg_row, h="left", sz=8)
    src_bg = C_LTRED if source=='Import' else C_LTGREEN
    src_color = C_RED if source=='Import' else C_TEAL
    write_cell(ws8, row, 6, source, bg=src_bg, color=src_color, bold=True, h="center", sz=8)
    write_cell(ws8, row, 7, item['uom'], bg=bg_row, h="center", sz=8)

    total_order = 0
    for ci, mk in enumerate(month_keys, start=8):
        v = item['orders'].get(mk, 0)
        total_order += v
        write_cell(ws8, row, ci, round(v) if v > 0 else 0, bg=bg_row, num_format=NUM_FMT,
                   bold=(v>0), color=C_NAVY if v>0 else C_GRAY3)
    write_cell(ws8, row, 12, round(total_order), bg=bg_row, bold=True, num_format=NUM_FMT,
               color=C_NAVY)

    row += 1
    sup_num += 1

# ═══════════════════════════════════════════════
# SHEET 9: Material Value Plan (Component Level)
# ═══════════════════════════════════════════════
print("Building Sheet 9: Material Value Plan - Component...")
ws9 = wb.create_sheet("Value Plan - Component")
ws9.sheet_view.showGridLines = False
ws9.freeze_panes = "E5"

ws9.merge_cells("A1:V1")
t = ws9["A1"]
t.value = "MATERIAL VALUE PLAN — COMPONENT LEVEL  (Values in IDR)"
t.font = font(bold=True, color=C_WHITE, size=13)
t.fill = fill(C_NAVY)
t.alignment = align(h="center")
ws9.row_dimensions[1].height = 30

ws9.merge_cells("A2:V2")
ws9["A2"].value = "Value = Quantity × Unit Price (IDR).  Consumption = Requirement value.  Receipt = Purchase value."
ws9["A2"].font = font(italic=True, color=C_NAVY, size=9)
ws9["A2"].fill = fill(C_LTBLUE)
ws9["A2"].alignment = align(h="center")

col_start = 5
for mk in month_keys:
    ws9.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_start+3)
    c = ws9.cell(row=3, column=col_start, value=mk)
    c.font = font(bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_NAVY)
    c.alignment = align(h="center"); c.border = thin_border()
    col_start += 4

for ci in range(1, 5):
    ws9.cell(row=3, column=ci).fill = fill(C_NAVY)
    ws9.cell(row=3, column=ci).border = thin_border()

sub9 = ['Comp Code','Description','Category','ABC']
for mk in month_keys:
    sub9 += ['Opening\n(IDR M)','Consumption\n(IDR M)','Receipt\n(IDR M)','Closing\n(IDR M)']
apply_subheader(ws9, 4, sub9, bg=C_NAVY)

ws9.column_dimensions['A'].width = 12
ws9.column_dimensions['B'].width = 35
ws9.column_dimensions['C'].width = 22
ws9.column_dimensions['D'].width = 7
for ci in range(5, 22):
    ws9.column_dimensions[get_column_letter(ci)].width = 13

IDR_M = 1_000_000  # millions
price_map = df_prices.set_index('Comp_Code')['Unit_Price'].to_dict()

row = 5
val_num = 1
prev_cat = None

for comp, months_data in sorted_comps:
    cat = months_data[0]['category']
    desc = comp_master.get(comp, {}).get('desc', 'Unknown')
    price = price_map.get(comp, 0)
    abc = abc_map.get(comp, '-')

    is_new_cat = cat != prev_cat
    if is_new_cat:
        ws9.merge_cells(start_row=row, start_column=1, end_row=row, end_column=20)
        cc = ws9.cell(row=row, column=1, value=f"  {cat}")
        cc.font = font(bold=True, color=C_WHITE, size=9)
        cc.fill = fill(C_NAVY)
        cc.alignment = align(h="left")
        ws9.row_dimensions[row].height = 14
        row += 1
        prev_cat = cat

    bg_row = C_WHITE if val_num % 2 == 0 else C_GRAY1
    abc_bg = {'A': C_LTRED, 'B': C_LTAMBER, 'C': C_LTGREEN}.get(str(abc), C_GRAY1)

    write_cell(ws9, row, 1, comp, bg=bg_row, h="center", sz=8)
    write_cell(ws9, row, 2, desc, bg=bg_row, h="left", sz=8)
    write_cell(ws9, row, 3, cat, bg=bg_row, h="left", sz=8)
    write_cell(ws9, row, 4, abc, bg=abc_bg, bold=True, h="center",
               color={'A': C_RED, 'B': C_AMBER, 'C': C_TEAL}.get(str(abc), C_BLACK))

    col = 5
    for md in months_data:
        op_val = (md['opening'] * price) / IDR_M
        cons_val = (md['requirement'] * price) / IDR_M
        rec_val = (md['receipt'] * price) / IDR_M
        cl_val = (md['closing'] * price) / IDR_M

        write_cell(ws9, row, col, round(op_val, 2), bg=bg_row, num_format="#,##0.00"); col+=1
        write_cell(ws9, row, col, round(cons_val, 2), bg=bg_row, num_format="#,##0.00"); col+=1
        write_cell(ws9, row, col, round(rec_val, 2), bg=bg_row if rec_val==0 else C_LTPURPLE,
                   num_format="#,##0.00", bold=(rec_val>0), color=C_PURPLE if rec_val>0 else C_BLACK); col+=1
        write_cell(ws9, row, col, round(cl_val, 2), bg=bg_row, num_format="#,##0.00"); col+=1

    row += 1
    val_num += 1

# ═══════════════════════════════════════════════
# SHEET 10: Material Value Plan (Category)
# ═══════════════════════════════════════════════
print("Building Sheet 10: Material Value Plan - Category...")
ws10 = wb.create_sheet("Value Plan - Category")
ws10.sheet_view.showGridLines = False
ws10.freeze_panes = "B5"

ws10.merge_cells("A1:R1")
t = ws10["A1"]
t.value = "MATERIAL VALUE PLAN — MATERIAL GROUP LEVEL  (Values in IDR Millions)"
t.font = font(bold=True, color=C_WHITE, size=13)
t.fill = fill(C_NAVY)
t.alignment = align(h="center")
ws10.row_dimensions[1].height = 30

ws10.merge_cells("A2:R2")
ws10["A2"].value = "Aggregated by Material Group  |  Values in IDR Millions"
ws10["A2"].font = font(italic=True, color=C_NAVY, size=9)
ws10["A2"].fill = fill(C_LTBLUE)
ws10["A2"].alignment = align(h="center")

col_start = 2
for mk in month_keys:
    ws10.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_start+3)
    c = ws10.cell(row=3, column=col_start, value=mk)
    c.font = font(bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_NAVY)
    c.alignment = align(h="center"); c.border = thin_border()
    col_start += 4
ws10.cell(row=3, column=1).fill = fill(C_NAVY)
ws10.cell(row=3, column=1).border = thin_border()

sub10 = ['Material Group']
for mk in month_keys:
    sub10 += ['Opening\n(IDR M)','Consumption\n(IDR M)','Receipt\n(IDR M)','Closing\n(IDR M)']
apply_subheader(ws10, 4, sub10, bg=C_NAVY)

ws10.column_dimensions['A'].width = 28
for ci in range(2, 19):
    ws10.column_dimensions[get_column_letter(ci)].width = 14

cat_val = {}
for comp, months_data in mat_plan.items():
    cat = months_data[0]['category']
    price = price_map.get(comp, 0)
    if cat not in cat_val:
        cat_val[cat] = {mk: {'op':0,'cons':0,'rec':0,'cl':0} for mk in month_keys}
    for md in months_data:
        mk = md['month']
        cat_val[cat][mk]['op'] += md['opening'] * price / IDR_M
        cat_val[cat][mk]['cons'] += md['requirement'] * price / IDR_M
        cat_val[cat][mk]['rec'] += md['receipt'] * price / IDR_M
        cat_val[cat][mk]['cl'] += md['closing'] * price / IDR_M

row = 5
grand10 = {mk: {'op':0,'cons':0,'rec':0,'cl':0} for mk in month_keys}
for i, (cat, data) in enumerate(sorted(cat_val.items())):
    bg_row = C_WHITE if i % 2 == 0 else C_GRAY1
    write_cell(ws10, row, 1, cat, bg=bg_row, bold=True, h="left", sz=9)
    col = 2
    for mk in month_keys:
        op=data[mk]['op']; cons=data[mk]['cons']; rec=data[mk]['rec']; cl=data[mk]['cl']
        write_cell(ws10, row, col, round(op,2), bg=bg_row, num_format="#,##0.00"); col+=1
        write_cell(ws10, row, col, round(cons,2), bg=bg_row, num_format="#,##0.00"); col+=1
        write_cell(ws10, row, col, round(rec,2), bg=C_LTPURPLE if rec>0 else bg_row,
                   bold=(rec>0), color=C_PURPLE if rec>0 else C_BLACK, num_format="#,##0.00"); col+=1
        write_cell(ws10, row, col, round(cl,2), bg=bg_row, num_format="#,##0.00"); col+=1
        grand10[mk]['op']+=op; grand10[mk]['cons']+=cons; grand10[mk]['rec']+=rec; grand10[mk]['cl']+=cl
    row += 1

write_cell(ws10, row, 1, "GRAND TOTAL", bg=C_NAVY, bold=True, color=C_WHITE, h="left", sz=10)
col = 2
for mk in month_keys:
    for key in ['op','cons','rec','cl']:
        v = grand10[mk][key]
        write_cell(ws10, row, col, round(v,2), bg=C_NAVY, bold=True, color=C_WHITE, num_format="#,##0.00"); col+=1

# ═══════════════════════════════════════════════
# SHEET 11: ABC Analysis
# ═══════════════════════════════════════════════
print("Building Sheet 11: ABC Analysis...")
ws11 = wb.create_sheet("ABC Analysis")
ws11.sheet_view.showGridLines = False

ws11.merge_cells("A1:M1")
t = ws11["A1"]
t.value = "ABC ANALYSIS — BOM COMPONENT ITEMS  (Value-wise Classification)"
t.font = font(bold=True, color=C_WHITE, size=13)
t.fill = fill(C_NAVY)
t.alignment = align(h="center")
ws11.row_dimensions[1].height = 30

ws11.merge_cells("A2:M2")
ws11["A2"].value = "A = High Value (prioritize), B = Medium Value, C = Low Value"
ws11["A2"].font = font(italic=True, color=C_NAVY, size=9)
ws11["A2"].fill = fill(C_LTBLUE)
ws11["A2"].alignment = align(h="center")

# ABC Summary first
ws11.merge_cells("A4:M4")
ws11["A4"].value = "ABC Summary — Annual Value Distribution"
ws11["A4"].font = font(bold=True, color=C_WHITE, size=10)
ws11["A4"].fill = fill(C_TEAL)
ws11["A4"].alignment = align(h="left")

sum_hdrs = ['ABC Class','# Items','% Items','Total Value (IDR M)','% Value','Focus Strategy']
for ci, h in enumerate(sum_hdrs, start=1):
    c = ws11.cell(row=5, column=ci, value=h)
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_BLUE)
    c.alignment = align(h="center"); c.border = thin_border()

# Build ABC data
abc_items = {}
for comp in mat_plan:
    abc = abc_map.get(comp, 'C')
    price = price_map.get(comp, 0)
    total_req = sum(mat_plan[comp][i]['requirement'] for i in range(4))
    total_val = total_req * price / IDR_M
    if abc not in abc_items:
        abc_items[abc] = {'count': 0, 'value': 0}
    abc_items[abc]['count'] += 1
    abc_items[abc]['value'] += total_val

total_items = sum(v['count'] for v in abc_items.values())
total_value = sum(v['value'] for v in abc_items.values())

abc_strategies = {
    'A': 'Daily monitoring, tight inventory control, dual sourcing, safety stock review weekly',
    'B': 'Weekly monitoring, standard reorder point, monthly inventory review',
    'C': 'Monthly monitoring, bulk ordering, min-max policy, lower priority'
}
abc_colors = {'A': C_RED, 'B': C_AMBER, 'C': C_TEAL}
abc_bgs = {'A': C_LTRED, 'B': C_LTAMBER, 'C': C_LTGREEN}

for ri, cls in enumerate(['A', 'B', 'C'], start=6):
    d = abc_items.get(cls, {'count': 0, 'value': 0})
    bg = abc_bgs[cls]
    write_cell(ws11, ri, 1, f"Class {cls}", bg=bg, bold=True, color=abc_colors[cls], h="center")
    write_cell(ws11, ri, 2, d['count'], bg=bg, h="center")
    pct_items = d['count']/total_items*100 if total_items > 0 else 0
    write_cell(ws11, ri, 3, round(pct_items,1), bg=bg, h="center", num_format="0.0\"%\"")
    write_cell(ws11, ri, 4, round(d['value'],2), bg=bg, h="right", num_format="#,##0.00")
    pct_val = d['value']/total_value*100 if total_value > 0 else 0
    write_cell(ws11, ri, 5, round(pct_val,1), bg=bg, h="center", num_format="0.0\"%\"")
    write_cell(ws11, ri, 6, abc_strategies.get(cls,''), bg=bg, h="left", sz=8)
    ws11.merge_cells(start_row=ri, start_column=6, end_row=ri, end_column=13)

# Detailed ABC list
ws11.merge_cells("A10:M10")
ws11["A10"].value = "Detailed ABC Item List — Sorted by Value (High to Low)"
ws11["A10"].font = font(bold=True, color=C_WHITE, size=10)
ws11["A10"].fill = fill(C_TEAL)
ws11["A10"].alignment = align(h="left")

abc_det_hdrs = ['#','ABC','Comp Code','Description','Category','Source','Lead Time',
                'Lot Size','Unit Price','4M Requirement','4M Value (IDR M)','% Cumulative','Action']
for ci, h in enumerate(abc_det_hdrs, start=1):
    c = ws11.cell(row=11, column=ci, value=h)
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_BLUE)
    c.alignment = align(h="center", wrap=True); c.border = thin_border()

ws11.column_dimensions['A'].width = 5
ws11.column_dimensions['B'].width = 7
ws11.column_dimensions['C'].width = 12
ws11.column_dimensions['D'].width = 35
ws11.column_dimensions['E'].width = 20
ws11.column_dimensions['F'].width = 10
ws11.column_dimensions['G'].width = 10
ws11.column_dimensions['H'].width = 12
ws11.column_dimensions['I'].width = 14
ws11.column_dimensions['J'].width = 14
ws11.column_dimensions['K'].width = 14
ws11.column_dimensions['L'].width = 12
ws11.column_dimensions['M'].width = 22

abc_detail_data = []
for comp in mat_plan:
    abc = abc_map.get(comp, 'C')
    price = price_map.get(comp, 0)
    desc = comp_master.get(comp, {}).get('desc', 'Unknown')
    cat = mat_plan[comp][0]['category']
    source = mat_plan[comp][0]['source']
    lt = mat_plan[comp][0]['lead_time']
    ls = mat_plan[comp][0]['lot_size']
    total_req = sum(mat_plan[comp][i]['requirement'] for i in range(4))
    total_val = total_req * price / IDR_M
    abc_detail_data.append({
        'abc': abc, 'comp': comp, 'desc': desc, 'cat': cat, 'source': source,
        'lt': lt, 'ls': ls, 'price': price, 'req': total_req, 'val': total_val
    })

abc_detail_data.sort(key=lambda x: -x['val'])
cum_val = 0
for ri, item in enumerate(abc_detail_data, start=12):
    cum_val += item['val']
    cum_pct = cum_val / total_value * 100 if total_value > 0 else 0
    abc = item['abc']
    bg = abc_bgs.get(abc, C_GRAY1)
    color = abc_colors.get(abc, C_BLACK)

    write_cell(ws11, ri, 1, ri-11, bg=bg, h="center", sz=8)
    write_cell(ws11, ri, 2, abc, bg=bg, bold=True, color=color, h="center")
    write_cell(ws11, ri, 3, item['comp'], bg=bg, h="center", sz=8)
    write_cell(ws11, ri, 4, item['desc'], bg=bg, h="left", sz=8)
    write_cell(ws11, ri, 5, item['cat'], bg=bg, h="left", sz=8)
    src_bg = C_LTRED if item['source']=='Import' else C_LTGREEN
    src_col = C_RED if item['source']=='Import' else C_TEAL
    write_cell(ws11, ri, 6, item['source'], bg=src_bg, color=src_col, bold=True, h="center", sz=8)
    write_cell(ws11, ri, 7, f"{int(item['lt'])}d", bg=bg, h="center", sz=8)
    write_cell(ws11, ri, 8, item['ls'], bg=bg, h="center", sz=8, num_format=NUM_FMT)
    write_cell(ws11, ri, 9, item['price'], bg=bg, h="right", sz=8, num_format=NUM_FMT)
    write_cell(ws11, ri, 10, round(item['req']), bg=bg, h="right", sz=8, num_format=NUM_FMT)
    write_cell(ws11, ri, 11, round(item['val'],2), bg=bg, h="right", sz=8, num_format="#,##0.00")
    write_cell(ws11, ri, 12, round(cum_pct,1), bg=bg, h="center", sz=8, num_format="0.0\"%\"")
    action = {'A':'Daily review', 'B':'Weekly review', 'C':'Monthly review'}.get(abc,'Review')
    write_cell(ws11, ri, 13, action, bg=bg, h="center", sz=8, color=color)

# ═══════════════════════════════════════════════
# SHEET 12: Dashboard
# ═══════════════════════════════════════════════
print("Building Sheet 12: Dashboard...")
ws12 = wb.create_sheet("📊 Dashboard", 0)  # put first
ws12.sheet_view.showGridLines = False

# Set all column widths
for ci in range(1, 25):
    ws12.column_dimensions[get_column_letter(ci)].width = 14
ws12.column_dimensions['A'].width = 3
ws12.column_dimensions['B'].width = 3

# Title
ws12.merge_cells("A1:X1")
t = ws12["A1"]
t.value = "INTEGRATED SUPPLY CHAIN PLANNING DASHBOARD"
t.font = font(bold=True, color=C_WHITE, size=16)
t.fill = fill(C_NAVY)
t.alignment = align(h="center")
ws12.row_dimensions[1].height = 40

ws12.merge_cells("A2:X2")
t2 = ws12["A2"]
t2.value = f"Plan Period: Jun-2025 to Sep-2025  |  {len(df_fc)} SKUs  |  {len(mat_plan)} Active Components"
t2.font = font(italic=True, color=C_NAVY, size=10)
t2.fill = fill(C_LTBLUE)
t2.alignment = align(h="center")
ws12.row_dimensions[2].height = 18

def dash_kpi(ws, row, col, label, value, unit="", bg=C_NAVY, val_color=C_AMBER):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
    ws.cell(row=row, column=col, value=label).font = font(bold=True, color=C_WHITE, size=8)
    ws.cell(row=row, column=col).fill = fill(bg)
    ws.cell(row=row, column=col).alignment = align(h="center")
    ws.cell(row=row, column=col).border = thin_border()
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+2)
    v_cell = ws.cell(row=row+1, column=col, value=value)
    v_cell.font = font(bold=True, color=val_color, size=16)
    v_cell.fill = fill(C_BLACK if bg==C_NAVY else C_GRAY1)
    v_cell.alignment = align(h="center")
    v_cell.border = thin_border()
    ws.row_dimensions[row+1].height = 28
    if unit:
        ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+2)
        u_cell = ws.cell(row=row+2, column=col, value=unit)
        u_cell.font = font(italic=True, color=C_GRAY3, size=8)
        u_cell.fill = fill(C_BLACK if bg==C_NAVY else C_GRAY1)
        u_cell.alignment = align(h="center")
        u_cell.border = thin_border()

# KPI Row
ws12.row_dimensions[4].height = 16
row = 4

# Label row
ws12.merge_cells("C4:X4")
ws12.cell(row=4, column=3, value="KEY PERFORMANCE INDICATORS").font = font(bold=True, color=C_WHITE, size=9)
ws12.cell(row=4, column=3).fill = fill(C_TEAL)
ws12.cell(row=4, column=3).alignment = align(h="center")

# KPI values
total_fc_4m = sum(sum(p['Forecast'][mk] for mk in month_keys) for p in prod_plan)
total_pr_4m = sum(sum(p['Production'][mk] for mk in month_keys) for p in prod_plan)
total_mat_rec = sum(sum(mat_plan[c][i]['receipt'] for i in range(4)) for c in mat_plan)
total_mat_val = sum(sum(mat_plan[c][i]['receipt'] * price_map.get(c,0) for i in range(4)) for c in mat_plan)
running_skus = len(df_fc[df_fc['SKU_Type']=='Running'])
new_skus = len(df_fc[df_fc['SKU_Type']=='New Product'])

kpis = [
    ("Total 4M Forecast", f"{total_fc_4m/1e6:.1f}M", "Units", C_NAVY, C_AMBER),
    ("Total 4M Production", f"{total_pr_4m/1e6:.1f}M", "Units", C_NAVY, C_LTGREEN),
    ("Active SKUs", len(df_fc), f"Running:{running_skus} | New:{new_skus}", C_TEAL, C_WHITE),
    ("Components Planned", len(mat_plan), "Active Materials", C_TEAL, C_WHITE),
    ("Total 4M Procurement", f"{total_mat_val/1e9:.1f}B", "IDR", C_NAVY, C_AMBER),
]

kpi_start_col = 3
for label, val, unit, bg, vc in kpis:
    dash_kpi(ws12, 5, kpi_start_col, label, val, unit, bg, vc)
    kpi_start_col += 4

ws12.row_dimensions[5].height = 16
ws12.row_dimensions[6].height = 28
ws12.row_dimensions[7].height = 16

# ── FG Inventory Days Cover by Category ──
r = 10
ws12.merge_cells(f"C{r}:X{r}")
ws12.cell(row=r, column=3, value="▶ FG CLOSING INVENTORY — DAYS COVER BY CATEGORY (Monthly)").font = font(bold=True, color=C_WHITE, size=10)
ws12.cell(row=r, column=3).fill = fill(C_TEAL)
ws12.cell(row=r, column=3).alignment = align(h="left")
ws12.cell(row=r, column=3).border = thin_border()
r += 1

ws12.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
c = ws12.cell(row=r, column=3, value='Category')
c.font = font(bold=True, color=C_WHITE, size=9); c.fill = fill(C_BLUE)
c.alignment = align(h="center"); c.border = thin_border()
for ci, h in enumerate(MONTHS, start=7):
    c = ws12.cell(row=r, column=ci, value=h)
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_BLUE)
    c.alignment = align(h="center"); c.border = thin_border()
r += 1

for i, (cat, data) in enumerate(sorted(cat_summary.items())):
    bg_row = C_WHITE if i % 2 == 0 else C_GRAY1
    ws12.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    write_cell(ws12, r, 3, cat, bg=bg_row, h="left", bold=True, sz=8)
    for ci, mk in enumerate(month_keys, start=7):
        fc = data[mk]['fc']; cl = data[mk]['cl']
        days = MONTH_DAYS[mk]
        dc = (cl/(fc/days)) if fc > 0 else 0
        dc_bg = C_LTGREEN if dc >= 10 else (C_LTAMBER if dc >= 5 else C_LTRED)
        c = ws12.cell(row=r, column=ci, value=round(dc,1))
        c.fill = fill(dc_bg)
        c.font = font(bold=True, color=C_TEAL if dc>=10 else (C_AMBER if dc>=5 else C_RED), size=9)
        c.alignment = align(h="center"); c.border = thin_border()
    r += 1

# ── BOM Component Inventory by Category ──
ws12.merge_cells(f"C{r+1}:X{r+1}")
ws12.cell(row=r+1, column=3, value="▶ BOM COMPONENT CLOSING INVENTORY — DAYS COVER BY MATERIAL GROUP (Monthly)").font = font(bold=True, color=C_WHITE, size=10)
ws12.cell(row=r+1, column=3).fill = fill(C_PURPLE)
ws12.cell(row=r+1, column=3).alignment = align(h="left")
ws12.cell(row=r+1, column=3).border = thin_border()
r += 2

ws12.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
c = ws12.cell(row=r, column=3, value='Material Group')
c.font = font(bold=True, color=C_WHITE, size=9); c.fill = fill(C_PURPLE)
c.alignment = align(h="center"); c.border = thin_border()
for ci, h in enumerate(MONTHS, start=7):
    c = ws12.cell(row=r, column=ci, value=h)
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_PURPLE)
    c.alignment = align(h="center"); c.border = thin_border()
r += 1

for i, (cat, data) in enumerate(sorted(cat_mat.items())):
    bg_row = C_WHITE if i % 2 == 0 else C_GRAY1
    ws12.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    write_cell(ws12, r, 3, cat, bg=bg_row, h="left", bold=True, sz=8)
    for ci, mk in enumerate(month_keys, start=7):
        req = data[mk]['req']; cl = data[mk]['cl']
        days = MONTH_DAYS[mk]
        dc = (cl/(req/days)) if req > 0 else 0
        dc_bg = C_LTGREEN if dc >= 15 else (C_LTAMBER if dc >= 7 else C_LTRED)
        c = ws12.cell(row=r, column=ci, value=round(dc,1))
        c.fill = fill(dc_bg)
        c.font = font(bold=True, color=C_TEAL if dc>=15 else (C_AMBER if dc>=7 else C_RED), size=9)
        c.alignment = align(h="center"); c.border = thin_border()
    r += 1

# ── Overall Summary ──
ws12.merge_cells(f"C{r+1}:X{r+1}")
ws12.cell(row=r+1, column=3, value="▶ OVERALL MONTHLY SUMMARY — FG vs MATERIAL INVENTORY DAYS COVER").font = font(bold=True, color=C_WHITE, size=10)
ws12.cell(row=r+1, column=3).fill = fill(C_NAVY)
ws12.cell(row=r+1, column=3).alignment = align(h="left")
ws12.cell(row=r+1, column=3).border = thin_border()
r += 2

ws12.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
c = ws12.cell(row=r, column=3, value='Metric')
c.font = font(bold=True, color=C_WHITE, size=9); c.fill = fill(C_NAVY)
c.alignment = align(h="center"); c.border = thin_border()
for ci, h in enumerate(MONTHS, start=7):
    c = ws12.cell(row=r, column=ci, value=h)
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_NAVY)
    c.alignment = align(h="center"); c.border = thin_border()
r += 1

overall_metrics = [
    ('FG Overall Days Cover', C_LTBLUE, C_BLUE),
    ('BOM Component Days Cover', C_LTPURPLE, C_PURPLE),
    ('Total FG Forecast (Units M)', C_GRAY1, C_NAVY),
    ('Total Production (Units M)', C_LTGREEN, C_TEAL),
]
for mi, (metric, bg, color) in enumerate(overall_metrics):
    ws12.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    write_cell(ws12, r, 3, metric, bg=bg, bold=True, color=color, h="left")
    for ci, mk in enumerate(month_keys, start=7):
        days = MONTH_DAYS[mk]
        if 'FG Overall' in metric:
            fc_tot = sum(p['Forecast'][mk] for p in prod_plan)
            cl_tot = sum(p['Closing'][mk] for p in prod_plan)
            val = (cl_tot/(fc_tot/days)) if fc_tot > 0 else 0
            v_fmt = DAY_FMT
        elif 'BOM Component' in metric:
            req_tot = sum(cat_mat[c][mk]['req'] for c in cat_mat)
            cl_tot = sum(cat_mat[c][mk]['cl'] for c in cat_mat)
            val = (cl_tot/(req_tot/days)) if req_tot > 0 else 0
            v_fmt = DAY_FMT
        elif 'Forecast' in metric:
            val = sum(p['Forecast'][mk] for p in prod_plan) / 1e6
            v_fmt = "#,##0.0"
        else:
            val = sum(p['Production'][mk] for p in prod_plan) / 1e6
            v_fmt = "#,##0.0"

        val_bg = C_LTGREEN if (metric in ['FG Overall Days Cover','BOM Component Days Cover'] and val >= 10) else \
                 (C_LTAMBER if (metric in ['FG Overall Days Cover','BOM Component Days Cover'] and val >= 5) else
                  (C_LTRED if metric in ['FG Overall Days Cover','BOM Component Days Cover'] else bg))
        c = ws12.cell(row=r, column=ci, value=round(val,1))
        c.fill = fill(val_bg)
        c.font = font(bold=True, color=color, size=9)
        c.alignment = align(h="center"); c.border = thin_border()
        c.number_format = v_fmt
    r += 1

# Legend
r += 1
ws12.merge_cells(f"C{r}:X{r}")
ws12.cell(row=r, column=3, value="LEGEND:").font = font(bold=True, color=C_NAVY, size=9)
ws12.cell(row=r, column=3).fill = fill(C_GRAY2)
ws12.cell(row=r, column=3).alignment = align(h="left")
r += 1

legends = [
    ("🟢 GREEN = Inventory at/above target", C_LTGREEN, C_TEAL),
    ("🟡 AMBER = Below target but within threshold", C_LTAMBER, C_AMBER),
    ("🔴 RED = Critical — below minimum threshold", C_LTRED, C_RED),
]
for ci_off, (leg, leg_bg, leg_col) in enumerate(legends):
    col_s = 3 + ci_off * 7
    ws12.merge_cells(start_row=r, start_column=col_s, end_row=r, end_column=col_s+6)
    lc = ws12.cell(row=r, column=col_s, value=leg)
    lc.fill = fill(leg_bg); lc.font = font(bold=True, color=leg_col, size=9)
    lc.alignment = align(h="center"); lc.border = thin_border()

# ═══════════════════════════════════════
# SAVE
# ═══════════════════════════════════════
output_path = OUTPUT_FILE
print(f"\nSaving to {output_path}...")
wb.save(output_path)
print("✓ Saved successfully!")

# Print sheet summary
print("\nSheets created:")
for i, s in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {s}")
